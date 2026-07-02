"""
Abgleich-Engine: Vergleicht die C16-Buchungsexport-Datei (Coor) mit der
eigenen Eingangsrechnungen-Excel und markiert Abweichungen gelb.

Annahmen (siehe README.md fuer Details):
- C16-Datei: flache Tabelle, Spalten u.a. 'Belegnr.', 'Belegnr.2', 'Istbetrag'.
- Eigene Datei: verschachtelte Bloecke. Kopfzeile pro Rechnung (Gewerke-Nr
  gefuellt) mit 'Ext. Re. Nr 1' und 'Ext. Re. Nr 2'. Darunter eine
  Unterueberschrift ('Nummer', 'Zahlung Netto', 'Zahlung Brutto') und
  danach 1..n Zahlungszeilen.
- SPALTEN WERDEN NICHT MEHR FEST ANGENOMMEN: Beide Dateien werden ueber
  ihre Spaltenueberschriften eingelesen (tolerant gegenueber Gross-/
  Kleinschreibung, Punkten, Bindestrichen und Leerzeichen). Die
  Kopfzeile darf auch weiter unten stehen (z.B. Logo/Titelzeilen davor).
  Dadurch funktioniert das Tool auch bei leicht anderem Aufbau
  (andere Objekte/Kostenstellen, eingefuegte oder umsortierte Spalten).
- Verknuepfung: Ext. Re. Nr 1/2 <-> Belegnr./Belegnr.2 (exakt, sowie ueber
  den Praefix vor dem ersten '/', da Ratenzahlungen in der C16 oft als
  'BASISNR/laufendenr' auftauchen).
- Abgleich je einzelner Zahlungszeile (nicht Summenebene): jeder
  Zahlung-Brutto (bzw. ersatzweise Netto) -Betrag muss unter dem
  zugehoerigen Beleg-Key in der C16 exakt (Toleranz 0.02) auftauchen.
- Bei Abweichung wird NUR die Betrag-Zelle (Netto/Brutto) der betroffenen
  Zahlungszeile gelb markiert.
- Automatische Korrektur erfolgt nur im eindeutigen Fall: genau eine
  Zahlungszeile im Rechnungsblock UND genau ein noch nicht verbrauchter
  C16-Betrag unter demselben Key. In allen anderen Faellen (mehrere
  Zahlungen, kein Key gefunden, mehrdeutig) wird NICHT automatisch
  korrigiert - die Zeile bleibt gelb markiert und im Log als
  "manuelle Pruefung noetig" aufgefuehrt.
"""

import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
TOLERANCE = 0.02  # Euro

# Wie viele Zeilen am Dateianfang nach der Kopfzeile durchsucht werden
# (falls z.B. Logo-/Titelzeilen vor der eigentlichen Tabelle stehen).
HEADER_SEARCH_ROWS = 30

# Obergrenze fuer automatische Korrekturen: Ein eindeutiger Fall (1 offene
# Zahlung + 1 freier C16-Betrag) wird nur automatisch korrigiert, wenn die
# Differenz zum urspruenglichen Betrag plausibel klein ist (Tippfehler,
# Rundung). Bei grossen Abweichungen (z.B. 6970 -> 300) handelt es sich
# vermutlich um zwei unterschiedliche Zahlungen, die zufaellig unter
# demselben Beleg-Key stehen - das darf NICHT automatisch verbucht werden,
# sondern muss als Abweichung gelb markiert und manuell geprueft werden.
AUTO_KORREKTUR_MAX_ABS_DIFF = 50.0     # Euro
AUTO_KORREKTUR_MAX_REL_DIFF = 0.20     # 20 %


def norm_key(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def prefix_key(v):
    """Praefix vor dem ersten '/' - fasst Ratenzahlungen wie
    'U18260100008/26-04' zu 'U18260100008' zusammen."""
    if v is None:
        return None
    s = str(v).strip()
    if "/" in s:
        return s.split("/")[0]
    return s


def to_float(v):
    """Wandelt einen Zellwert tolerant in float um. Versteht auch als Text
    gespeicherte Betraege in deutschem ('1.000,00') und englischem
    ('1000.00') Format. Gibt None zurueck, wenn kein Betrag lesbar ist
    (statt zu crashen)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").replace("EUR", "").strip()
    if not s:
        return None
    if "," in s and "." in s:
        # deutsches Format: Punkt = Tausender, Komma = Dezimal
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    elif re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
        # nur Punkte in 3er-Gruppen ohne Komma: deutscher Tausenderpunkt
        # ('1.500' = 1500, nicht 1,5)
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def _norm_header(v):
    """Normalisiert eine Spaltenueberschrift fuer den toleranten Vergleich:
    Kleinbuchstaben, ohne Punkte/Bindestriche/Unterstriche/Leerzeichen.
    'Ext. Re. Nr 1', 'Ext.Re.Nr. 1' und 'ext re nr 1' werden damit gleich."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = re.sub(r"[.\-_/\s]+", "", s)
    return s


def _header_map(cells):
    """Baut {normalisierte Ueberschrift: Spaltennummer} aus einer Zeile."""
    m = {}
    for i, c in enumerate(cells):
        h = _norm_header(c)
        if h and h not in m:  # bei Dubletten gewinnt die erste Spalte
            m[h] = i + 1
    return m


def _find_col(hmap, *candidates, contains=None):
    """Sucht eine Spalte anhand mehrerer moeglicher (normalisierter)
    Ueberschriften; optional zusaetzlich per Teilstring-Suche."""
    for cand in candidates:
        if cand in hmap:
            return hmap[cand]
    if contains:
        for h, col in hmap.items():
            if contains in h:
                return col
    return None


# ---------------------------------------------------------------------------
# C16 (Coor-Export)
# ---------------------------------------------------------------------------

@dataclass
class C16Row:
    row: int
    belegnr: str
    belegnr2: str
    betrag: float
    kreditor: str = None


def _find_c16_header(ws):
    """Sucht in den ersten HEADER_SEARCH_ROWS Zeilen die Kopfzeile mit
    'Belegnr.' und 'Istbetrag' (tolerant). Gibt (header_row, hmap) oder
    (None, None) zurueck."""
    for r in range(1, min(HEADER_SEARCH_ROWS, ws.max_row) + 1):
        hmap = _header_map([c.value for c in ws[r]])
        if _find_col(hmap, "belegnr") and _find_col(hmap, "istbetrag"):
            return r, hmap
    return None, None


def load_c16(path):
    """Laedt die C16-Datei. Sucht in allen Sheets nach der Kopfzeile mit
    'Belegnr.' und 'Istbetrag' - egal in welcher Zeile/Spalte sie steht."""
    wb = openpyxl.load_workbook(path, data_only=True)
    candidates = []  # (sheet, header_row, hmap) - alle Sheets mit passender Kopfzeile
    for sheet in wb.worksheets:
        hr, hm = _find_c16_header(sheet)
        if hr is not None:
            candidates.append((sheet, hr, hm))
    if not candidates:
        raise ValueError(
            "In der C16-Datei wurde keine Kopfzeile mit den Spalten "
            "'Belegnr.' und 'Istbetrag' gefunden. Ist das wirklich die "
            "C16-Exportdatei?"
        )

    def _read_rows(ws, header_row, hmap):
        col_beleg = _find_col(hmap, "belegnr")
        col_beleg2 = _find_col(hmap, "belegnr2")
        col_betrag = _find_col(hmap, "istbetrag")
        col_kreditor = _find_col(hmap, "debitorkreditorname", contains="kreditor")
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            belegnr = norm_key(ws.cell(r, col_beleg).value)
            belegnr2 = norm_key(ws.cell(r, col_beleg2).value) if col_beleg2 else None
            betrag = to_float(ws.cell(r, col_betrag).value)
            kreditor = ws.cell(r, col_kreditor).value if col_kreditor else None
            if belegnr is None and belegnr2 is None and betrag is None:
                continue
            rows.append(C16Row(row=r, belegnr=belegnr, belegnr2=belegnr2, betrag=betrag, kreditor=kreditor))
        return rows

    # Falls mehrere Sheets eine passende Kopfzeile haben (z.B. Helper-/
    # Legende-Sheets), gewinnt das erste Sheet, das auch Datenzeilen hat.
    best_rows = None
    for ws, hr, hm in candidates:
        rows = _read_rows(ws, hr, hm)
        if rows:
            return rows
        if best_rows is None:
            best_rows = rows
    return best_rows


def build_indexes(c16_rows):
    exact_idx = {}
    prefix_idx = {}
    for row in c16_rows:
        for k in (row.belegnr, row.belegnr2):
            if not k:
                continue
            exact_idx.setdefault(k, []).append(row)
            pk = prefix_key(k)
            prefix_idx.setdefault(pk, []).append(row)
    return exact_idx, prefix_idx


# ---------------------------------------------------------------------------
# Eigene Eingangsrechnungen-Datei
# ---------------------------------------------------------------------------

@dataclass
class PaymentLine:
    row: int
    nummer: str
    netto: float
    brutto: float
    status: str = "OFFEN"          # OK / ABWEICHUNG / KEIN_BELEG / KORRIGIERT
    matched_betrag: float = None
    note: str = ""


@dataclass
class InvoiceBlock:
    header_row: int
    auftrag: str
    beschreibung: str
    ext1: str
    ext2: str
    payments: list = field(default_factory=list)


@dataclass
class OwnLayout:
    """Automatisch erkanntes Spaltenlayout der eigenen Datei."""
    header_row: int
    col_gewerk: int
    col_auftrag: int
    col_beschreibung: int
    col_ext1: int
    col_ext2: int      # kann None sein (Spalte fehlt in manchen Vorlagen)
    col_nummer: int
    col_netto: int
    col_brutto: int

    def beschreibung_text(self):
        def L(c):
            return openpyxl.utils.get_column_letter(c) if c else "-"
        return (
            f"Kopfzeile in Zeile {self.header_row}; Gewerke-Nr={L(self.col_gewerk)}, "
            f"Ext.Re.Nr1={L(self.col_ext1)}, Ext.Re.Nr2={L(self.col_ext2)}, "
            f"Nummer={L(self.col_nummer)}, Netto={L(self.col_netto)}, Brutto={L(self.col_brutto)}"
        )


def _find_own_header(ws):
    """Sucht die Haupt-Kopfzeile der eigenen Datei ('Gewerke-Nr' +
    'Ext. Re. Nr 1'). Gibt (header_row, hmap) oder (None, None) zurueck."""
    for r in range(1, min(HEADER_SEARCH_ROWS, ws.max_row) + 1):
        hmap = _header_map([c.value for c in ws[r]])
        if _find_col(hmap, "gewerkenr") and _find_col(hmap, "extrenr1"):
            return r, hmap
    return None, None


def _find_payment_subheader(ws, start_row):
    """Sucht unterhalb der Kopfzeile die erste Unterueberschrift der
    Zahlungszeilen ('Nummer' + 'Zahlung Netto'/'Zahlung Brutto') und gibt
    deren Spalten zurueck: (col_nummer, col_netto, col_brutto)."""
    for r in range(start_row + 1, min(start_row + 200, ws.max_row) + 1):
        hmap = _header_map([c.value for c in ws[r]])
        col_nummer = _find_col(hmap, "nummer")
        col_netto = _find_col(hmap, "zahlungnetto")
        col_brutto = _find_col(hmap, "zahlungbrutto")
        if col_nummer and (col_netto or col_brutto):
            return col_nummer, col_netto, col_brutto
    return None, None, None


def detect_own_layout(wb):
    """Erkennt Sheet + Spaltenlayout der eigenen Datei anhand der
    Ueberschriften. Wirft einen verstaendlichen Fehler, wenn die
    Pflichtspalten nicht gefunden werden (lieber klarer Fehler als
    stillschweigend falsche Spalten lesen)."""
    found_main_header = False
    ws = header_row = hmap = None
    col_nummer = col_netto = col_brutto = None
    for sheet in wb.worksheets:
        hr, hm = _find_own_header(sheet)
        if hr is None:
            continue
        found_main_header = True
        cn, cnet, cbru = _find_payment_subheader(sheet, hr)
        if cn and (cnet or cbru):
            # erstes Sheet, das Kopfzeile UND Zahlungs-Unterueberschrift
            # hat, gewinnt (Helper-/Legende-Sheets fallen so durch)
            ws, header_row, hmap = sheet, hr, hm
            col_nummer, col_netto, col_brutto = cn, cnet, cbru
            break
    if ws is None:
        if found_main_header:
            raise ValueError(
                "Die Unterueberschrift der Zahlungszeilen ('Nummer', "
                "'Zahlung Netto', 'Zahlung Brutto') wurde nicht gefunden. "
                "Der Aufbau der Datei weicht zu stark von der bekannten "
                "Vorlage ab - bitte pruefen."
            )
        raise ValueError(
            "In der eigenen Datei wurde keine Kopfzeile mit den Spalten "
            "'Gewerke-Nr' und 'Ext. Re. Nr 1' gefunden. Ist das wirklich "
            "die Eingangsrechnungen-Datei?"
        )
    col_gewerk = _find_col(hmap, "gewerkenr")
    col_auftrag = _find_col(hmap, "auftragsnummer", contains="auftragsnummer")
    col_beschreibung = _find_col(hmap, "hauptauftrag", contains="hauptauftrag")
    col_ext1 = _find_col(hmap, "extrenr1")
    col_ext2 = _find_col(hmap, "extrenr2")

    layout = OwnLayout(
        header_row=header_row,
        col_gewerk=col_gewerk,
        col_auftrag=col_auftrag or col_gewerk + 1,
        col_beschreibung=col_beschreibung or (col_auftrag or col_gewerk + 1) + 1,
        col_ext1=col_ext1,
        col_ext2=col_ext2,
        col_nummer=col_nummer,
        col_netto=col_netto or col_brutto,
        col_brutto=col_brutto or col_netto,
    )
    return ws, layout


def load_own_blocks(path):
    """Parst die verschachtelte Struktur der eigenen Datei in
    InvoiceBlock-Objekte mit ihren PaymentLine-Kindern.
    Gibt (wb, ws, blocks, layout) zurueck."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws, layout = detect_own_layout(wb)
    blocks = []
    current = None
    r = layout.header_row + 1
    max_row = ws.max_row
    while r <= max_row:
        gewerk = norm_key(ws.cell(r, layout.col_gewerk).value)  # norm_key: Whitespace-Zellen zaehlen als leer
        nummer_val = ws.cell(r, layout.col_nummer).value
        if gewerk is not None:
            current = InvoiceBlock(
                header_row=r,
                auftrag=ws.cell(r, layout.col_auftrag).value,
                beschreibung=ws.cell(r, layout.col_beschreibung).value,
                ext1=norm_key(ws.cell(r, layout.col_ext1).value),
                ext2=norm_key(ws.cell(r, layout.col_ext2).value) if layout.col_ext2 else None,
            )
            blocks.append(current)
        elif _norm_header(nummer_val) == "nummer":
            pass  # Unterueberschrift, keine Daten
        elif norm_key(nummer_val) is not None:
            netto = ws.cell(r, layout.col_netto).value
            brutto = ws.cell(r, layout.col_brutto).value
            # Nur als Zahlungszeile werten, wenn sie auch wie eine aussieht:
            # Zahlungs-Nummer ist rein numerisch (z.B. '00025') ODER
            # mindestens ein Betrag ist als Zahl lesbar. Sonst handelt es
            # sich um eine Struktur-/Fortsetzungszeile des Rechnungskopfs
            # (z.B. 'Notare Haubold & Haubold, RE21/01008' im Betragsfeld),
            # die frueher faelschlich als Zahlung gelesen wurde und im ganz
            # alten Code sogar zum Absturz fuehrte.
            looks_like_payment = (
                to_float(netto) is not None
                or to_float(brutto) is not None
                or re.fullmatch(r"\d+", str(nummer_val).strip()) is not None
            )
            if current is not None and looks_like_payment:
                current.payments.append(
                    PaymentLine(row=r, nummer=nummer_val, netto=netto, brutto=brutto)
                )
            else:
                _backfill_header(ws, layout, current, r)
        else:
            # Gewerke- und Nummer-Spalte beide leer: kommt in der Vorlage
            # manchmal vor, wenn die Kopfzeile eines Blocks ueber zwei
            # physische Zeilen verteilt ist (Ext.Re.Nr1/Re.Datum/Re.Eingang
            # landen dann nicht in derselben Zeile wie Gewerke-Nr/
            # Auftragsnummer). In dem Fall die fehlenden Kopf-Werte in den
            # aktuellen Block nachtragen, statt sie faelschlich als
            # Zahlungszeile zu lesen.
            if current is not None and current.ext1 is None:
                val = ws.cell(r, layout.col_ext1).value
                if val is not None and not isinstance(val, (int, float)):
                    current.ext1 = norm_key(val)
            if current is not None and current.ext2 is None and layout.col_ext2:
                val = ws.cell(r, layout.col_ext2).value
                if val is not None and not isinstance(val, (int, float)):
                    current.ext2 = norm_key(val)
        r += 1
    return wb, ws, blocks, layout


# ---------------------------------------------------------------------------
# Abgleich
# ---------------------------------------------------------------------------

def compare(c16_rows, blocks):
    """Fuehrt den eigentlichen Abgleich durch und setzt status/matched_betrag/
    note auf jeder PaymentLine. Gibt eine Zusammenfassung (dict) zurueck."""
    exact_idx, prefix_idx = build_indexes(c16_rows)

    summary = {
        "bloecke": len(blocks),
        "zahlungen": sum(len(b.payments) for b in blocks),
        "ok": 0,
        "abweichung": 0,
        "kein_beleg": 0,
        "ohne_betrag": 0,
        "auto_korrigiert": 0,
        "manuelle_pruefung": 0,
    }

    # C16-Zeilen werden GLOBAL verbraucht (ueber alle Bloecke hinweg), damit
    # dieselbe C16-Buchung nicht in zwei Rechnungsbloecken gleichzeitig als
    # "OK" verbucht werden kann (z.B. bei gleicher Ext.Re.Nr oder gleichem
    # Ratenzahlungs-Praefix in mehreren Bloecken).
    used_c16_rows = set()

    for b in blocks:
        pool = []
        seen = set()
        for k in (b.ext1, b.ext2):
            if not k:
                continue
            for src in (exact_idx.get(k, []), prefix_idx.get(prefix_key(k), [])):
                for c in src:
                    if c.row not in seen:
                        pool.append(c)
                        seen.add(c.row)

        unmatched_payments = []
        for p in b.payments:
            amt = to_float(p.brutto if p.brutto is not None else p.netto)
            if amt is None:
                # Kein lesbarer Betrag in der Zahlungszeile: nicht lautlos
                # ueberspringen, sondern zur manuellen Pruefung markieren.
                p.status = "OHNE_BETRAG"
                p.note = "Kein (lesbarer) Betrag in der Zahlungszeile eingetragen"
                summary["ohne_betrag"] += 1
                continue
            if not pool:
                p.status = "KEIN_BELEG"
                p.note = "Kein Beleg mit dieser Ext.Re.Nr in der C16 gefunden"
                summary["kein_beleg"] += 1
                unmatched_payments.append(p)
                continue
            found = None
            for c in pool:
                if c.row in used_c16_rows:
                    continue
                if c.betrag is not None and abs(float(c.betrag) - amt) <= TOLERANCE:
                    found = c
                    break
            if found is not None:
                used_c16_rows.add(found.row)
                p.status = "OK"
                p.matched_betrag = found.betrag
                summary["ok"] += 1
            else:
                p.status = "ABWEICHUNG"
                p.note = "Betrag nicht in C16 unter diesem Beleg gefunden"
                summary["abweichung"] += 1
                unmatched_payments.append(p)

        # Eindeutige Auto-Korrektur: genau 1 offene Zahlung + genau 1 freier C16-Betrag,
        # UND die Differenz ist plausibel klein (siehe AUTO_KORREKTUR_MAX_*).
        free = [c for c in pool if c.row not in used_c16_rows]
        auto_korrigiert = False
        if len(unmatched_payments) == 1 and len(free) == 1 and free[0].betrag is not None:
            p = unmatched_payments[0]
            old = to_float(p.brutto if p.brutto is not None else p.netto)
            new = free[0].betrag
            diff = abs(float(new) - float(old))
            rel = diff / abs(float(old)) if old else float("inf")
            if diff <= AUTO_KORREKTUR_MAX_ABS_DIFF or rel <= AUTO_KORREKTUR_MAX_REL_DIFF:
                used_c16_rows.add(free[0].row)
                was_kein_beleg = p.status == "KEIN_BELEG"
                p.status = "KORRIGIERT"
                p.matched_betrag = new
                p.note = f"Automatisch korrigiert: {old} -> {new} (C16 Zeile {free[0].row})"
                summary["auto_korrigiert"] += 1
                if was_kein_beleg:
                    summary["kein_beleg"] -= 1
                else:
                    summary["abweichung"] -= 1
                auto_korrigiert = True
            else:
                p.note = (
                    f"{p.note} (1 moeglicher C16-Betrag {new} gefunden, aber Differenz "
                    f"zu {old} ist zu gross fuer automatische Korrektur - bitte manuell pruefen)"
                )
        if not auto_korrigiert:
            for p in unmatched_payments:
                if p.status in ("ABWEICHUNG", "KEIN_BELEG"):
                    summary["manuelle_pruefung"] += 1

    return summary


# ---------------------------------------------------------------------------
# Ausgabedateien
# ---------------------------------------------------------------------------

def _mark_payment(ws, p, betrag_cols, note):
    """Markiert die Betrag-Zellen einer Zahlungszeile gelb. Sind beide
    Betrag-Zellen leer (Fall OHNE_BETRAG), wird trotzdem die letzte
    Betrag-Spalte markiert, damit die Zeile sichtbar auffaellt."""
    filled = 0
    for col in betrag_cols:
        cell = ws.cell(p.row, col)
        if cell.value is not None:
            cell.fill = YELLOW_FILL
            cell.comment = Comment(note, "Abgleich-Tool")
            filled += 1
    if filled == 0:
        cell = ws.cell(p.row, betrag_cols[-1])
        cell.fill = YELLOW_FILL
        cell.comment = Comment(note, "Abgleich-Tool")


def _unhide_row(ws, row):
    """Klappt eine Zeile auf, falls sie Teil einer eingeklappten
    Excel-Gruppierung (Gliederung) ist. Ohne das waeren gelb markierte
    Zahlungszeilen fuer den Nutzer unsichtbar, weil die Vorlage die
    Zahlungszeilen standardmaessig eingeklappt/versteckt anzeigt."""
    ws.row_dimensions[row].hidden = False


def write_marked_copy(src_path, out_path, ws_title, blocks, layout):
    """Schreibt eine Kopie der Originaldatei, in der die Betrag-Zellen
    abweichender Zahlungszeilen gelb markiert sind (Original bleibt
    unveraendert erhalten)."""
    wb = openpyxl.load_workbook(src_path)  # ohne data_only, um Formeln/Format zu erhalten
    ws = wb[ws_title]
    betrag_cols = sorted({layout.col_netto, layout.col_brutto})
    for b in blocks:
        for p in b.payments:
            if p.status in ("ABWEICHUNG", "KEIN_BELEG", "OHNE_BETRAG"):
                _mark_payment(ws, p, betrag_cols, p.note)
                _unhide_row(ws, p.row)
    wb.save(out_path)


def write_corrected_copy(src_path, out_path, ws_title, blocks, layout):
    """Schreibt eine Kopie, in der eindeutig korrigierbare Betraege
    ueberschrieben wurden. Alles, was manuelle Pruefung braucht, bleibt
    unveraendert, aber weiterhin gelb markiert."""
    wb = openpyxl.load_workbook(src_path)
    ws = wb[ws_title]
    betrag_cols = sorted({layout.col_netto, layout.col_brutto})
    for b in blocks:
        for p in b.payments:
            if p.status == "KORRIGIERT":
                if p.brutto is not None:
                    ws.cell(p.row, layout.col_brutto).value = p.matched_betrag
                if p.netto is not None and p.brutto is None:
                    ws.cell(p.row, layout.col_netto).value = p.matched_betrag
                cell = ws.cell(p.row, layout.col_brutto if p.brutto is not None else layout.col_netto)
                cell.comment = Comment(p.note, "Abgleich-Tool")
                _unhide_row(ws, p.row)
            elif p.status in ("ABWEICHUNG", "KEIN_BELEG", "OHNE_BETRAG"):
                _mark_payment(ws, p, betrag_cols, p.note + " (manuelle Pruefung noetig)")
                _unhide_row(ws, p.row)
    wb.save(out_path)


def run_abgleich(c16_path, own_path, out_dir):
    """High-level Einstiegspunkt fuer die GUI. Gibt (summary, marked_path,
    corrected_path, log_lines) zurueck."""
    import os

    c16_rows = load_c16(c16_path)
    wb_own, ws_own, blocks, layout = load_own_blocks(own_path)
    summary = compare(c16_rows, blocks)

    base = os.path.splitext(os.path.basename(own_path))[0]
    marked_path = os.path.join(out_dir, f"{base}_markiert.xlsx")
    corrected_path = os.path.join(out_dir, f"{base}_korrigiert.xlsx")

    write_marked_copy(own_path, marked_path, ws_own.title, blocks, layout)
    write_corrected_copy(own_path, corrected_path, ws_own.title, blocks, layout)

    log_lines = [
        f"Erkanntes Layout: {layout.beschreibung_text()}",
        f"Rechnungsbloecke geprueft: {summary['bloecke']}",
        f"Zahlungszeilen geprueft: {summary['zahlungen']}",
        f"  OK (uebereinstimmend): {summary['ok']}",
        f"  Automatisch korrigiert: {summary['auto_korrigiert']}",
        f"  Abweichung (Betrag nicht gefunden): {summary['abweichung']}",
        f"  Kein passender Beleg in C16: {summary['kein_beleg']}",
        f"  Zahlungszeile ohne (lesbaren) Betrag: {summary['ohne_betrag']}",
        f"  -> insgesamt gelb markiert / manuelle Pruefung: "
        f"{summary['abweichung'] + summary['kein_beleg'] + summary['ohne_betrag']}",
    ]
    return summary, marked_path, corrected_path, log_lines
